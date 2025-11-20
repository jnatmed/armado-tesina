# Utils.py
import numpy as np
import pandas as pd
import json
import os, re
import hashlib
import datetime
from typing import Optional, Dict, Any
# Para el formato condicional en Excel
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule


import traceback
from pathlib import Path

_ILLEGAL_NAME = re.compile(r'[<>:"\\|?*\x00]')

class Utils:
    # --------------------- Utilidades ---------------------
   

    def _loggable_random_state(self):
        if isinstance(self._seed_init, (int, np.integer)):
            return int(self._seed_init)
        if self._seed_init is None:
            return None
        return str(self._seed_init)

    def _motivo(self, code, detail=None):
        """
        Normaliza motivos de corte.
        """
        if detail is not None:
            return f"{code}:{detail}"
        return code

    def _push_trace(self, **kv):
        """
        Agrega un evento a self.debug_trace con timestamp.
        """
        kv = dict(kv)
        kv.setdefault("ts", pd.Timestamp.now().isoformat())
        self.debug_trace.append(kv)

    def reset_logs(self):
        """Resetea logs por CLASE, por MUESTRA y metadatos."""
        self.logs_por_clase = []
        self.logs_por_muestra = []
        self.meta_experimento = {}
        self._meta = {}
        # diagnósticos (solo los que quedan vigentes con LSD)
        self._diag_densidad = None

    @staticmethod
    def tag_p(criterio: str) -> str:
        return "P" + str(criterio).lower().strip()

    @staticmethod
    def safe_token(s: str) -> str:
        """Sanitiza un fragmento para NOMBRE DE ARCHIVO (no toca la ruta)."""
        s = str(s).replace('\r','').replace('\n','').replace('\t','').strip()
        s = s.rstrip(' .')                          # sin espacio/punto de cierre
        return re.sub(r'[^A-Za-z0-9._-]+', '', s)  # solo seguro en Windows

    def atomic_save_csv_strict(self, df: pd.DataFrame, path_str: str, fallback_dir="C:/datasets") -> str:
        """Crea carpeta, sanea filename, maneja long-path y guarda atómico."""
        p = Path(path_str)
        # saneo SOLO el filename
        name = Utils.safe_token(p.name)
        if re.search(r'[<>:"\\|?*\x00]', name):
            raise ValueError(f"Filename inválido tras sanitizar: {repr(name)}")
        p = p.with_name(name)

        # long-path fallback si hiciera falta
        abs_guess = p.resolve(strict=False)
        if len(str(abs_guess)) > 240:  # margen para Windows sin long paths
            base = Path(fallback_dir); base.mkdir(parents=True, exist_ok=True)
            p = base / p.name
            print(f"↪ Ruta larga: guardo en {p}")

        p.parent.mkdir(parents=True, exist_ok=True)
        tmp = p.with_suffix(p.suffix + ".tmp")
        df.to_csv(tmp, index=False)
        os.replace(tmp, p)
        return str(p)


    def _debug_path(self, path_in) -> None:
        p = Path(path_in)
        print("=== DEBUG PATH ===")
        print("repr(path):", repr(str(path_in)))
        print("cwd:", os.getcwd())
        try:
            abs_guess = p.resolve(strict=False)
            print("abs path:", str(abs_guess))
            print("abs len:", len(str(abs_guess)))
        except Exception as e:
            print("resolve() lanzó:", type(e).__name__, e)

        print("parent:", str(p.parent))
        print("parent exists?", p.parent.exists(), "is_dir?", p.parent.is_dir())

        # ✅ Chequear SOLO el nombre de archivo (sin '/')
        name = p.name
        print("filename:", repr(name))
        ilegales = re.findall(r'[<>:"\\|?*\x00]', name)
        if ilegales:
            print("❗ illegal char(s) en filename:", ilegales)
        if name.endswith(' ') or name.endswith('.'):
            print("❗ filename termina con espacio o punto")

        # ¿parent escribible?
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            probe = p.parent / "_probe_write.tmp"
            with open(probe, "wb") as f:
                f.write(b"ok")
            os.replace(probe, p.parent / "_probe_write.ok")
            print("✅ Escritura en parent OK")
        except Exception as e:
            print("❌ NO se pudo escribir en parent:", type(e).__name__, e)
            traceback.print_exc()
        print("==================")

    def exportar_log_csv(self, path_salida):
        """Exporta el log por CLASE con diagnóstico."""
        if not self.logs_por_clase:
            print("⚠️ No hay log POR CLASE para exportar. CSV")
            return

        p = Path(path_salida)
        # self._debug_path(p)  # DEBUG
        p.parent.mkdir(parents=True, exist_ok=True)
        tmp = p.with_suffix(p.suffix + ".tmp")
        pd.DataFrame(self.logs_por_clase).to_csv(tmp, index=False)
        os.replace(tmp, p)
        print(f"📁 Log por clase guardado en: {p}")

    def exportar_log_muestras_excel_debug(self, path_salida,
                                        resaltar_no_filtradas=True,
                                        indices_resaltar=None):
        """
        Versión instrumentada para aislar fallos de escritura Excel.
        """
        print("==== DEBUG exportar_log_muestras_excel ====")
        print("→ path_salida repr:", repr(path_salida))
        print("→ cwd:", os.getcwd())

        try:
            # -- Comprobación de estructura
            p = Path(path_salida)
            print("→ absolute:", str(p.resolve(strict=False)))
            print("→ exists parent:", p.parent.exists(), "is_dir:", p.parent.is_dir())
            print("→ filename:", p.name, "len(abs):", len(str(p.resolve(strict=False))))
            for ch in '<>:"/\\|?*\x00':
                if ch in path_salida:
                    print("⚠️ Illegal char found:", repr(ch))

            # -- Datos
            if not self.logs_por_muestra:
                print("⚠️ No hay logs.")
                return
            df = pd.DataFrame(self.logs_por_muestra)

            print("→ DataFrame shape:", df.shape)
            print("→ Columns:", list(df.columns))

            # -- Serialización
            cols_json = ("vecinos_all", "clase_vecinos_all", "dist_all",
                        "vecinos_min", "dist_vecinos_min")
            for col in cols_json:
                if col in df.columns:
                    df[col] = df[col].apply(
                        lambda v: json.dumps(v, ensure_ascii=False)
                        if isinstance(v, (list, tuple, np.ndarray))
                        else v
                    )

            # -- Escritura segura
            print("→ Intentando abrir ExcelWriter...")
            try:
                p.parent.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                print("❌ mkdir falló:", type(e).__name__, e)
                traceback.print_exc()

            try:
                with pd.ExcelWriter(str(p), engine="openpyxl") as writer:
                    hoja = "Log_Muestras"
                    df.to_excel(writer, sheet_name=hoja, index=False)
                print("✅ Escritura exitosa en:", p)
            except Exception as e:
                print("❌ Error al escribir Excel:", type(e).__name__, e)
                traceback.print_exc()
        except Exception as e_outer:
            print("❌ EXCEPCIÓN GENERAL:", type(e_outer).__name__, e_outer)
            traceback.print_exc()
        finally:
            print("==== FIN DEBUG ====")


    def exportar_log_muestras_excel(
        self, 
        path_salida,
        resaltar_no_filtradas: bool = True,
        indices_resaltar=None,
        append: bool = True
    ):
        """
        Exporta el log POR MUESTRA a un Excel.

        Parámetros:
        -----------
        append : bool (default=True)
            True  → si el archivo ya existe, agrega filas nuevas.
            False → ignora el archivo si existe y sobrescribe con los logs actuales.

        NOTA:
        - Se mantiene la lógica de acumulación por dataset del PCSMOTE.
        - El flag "append" solo controla el archivo externo final.
        """

        if not getattr(self, "logs_por_muestra", None):
            print("⚠️ No hay log POR MUESTRA para exportar.")
            return

        # DataFrame base con los logs de ESTA ejecución
        df = pd.DataFrame(self.logs_por_muestra)

        # ───────────── Columna extra: 'configuracion' ─────────────
        nombre_configuracion = getattr(self, "nombre_configuracion", None)
        if nombre_configuracion is not None:
            if "configuracion" not in df.columns:
                if "dataset" in df.columns:
                    idx_dataset = list(df.columns).index("dataset")
                    pos_insert = idx_dataset + 1
                else:
                    pos_insert = 0
                df.insert(pos_insert, "configuracion", str(nombre_configuracion))

        # Serializar listas/arrays a JSON para que Excel no falle
        cols_json = ("vecinos_all", "clase_vecinos_all", "dist_all",
                    "vecinos_min", "dist_vecinos_min")
        for col in cols_json:
            if col in df.columns:
                df[col] = df[col].apply(
                    lambda v: json.dumps(v, ensure_ascii=False)
                    if isinstance(v, (list, tuple, np.ndarray))
                    else v
                )

        # ───────────── Acumulación por dataset (solo para PCSMOTE) ─────────────
        df_salida = df  # por defecto, sin acumulación interna

        nombre_dataset = getattr(self, "nombre_dataset", None)
        cls = type(self)

        if hasattr(cls, "acumular_logs_por_muestra_por_dataset") and \
        hasattr(cls, "obtener_logs_por_muestra_acumulados") and \
        nombre_dataset is not None:

            # 1) acumular este bloque en memoria
            cls.acumular_logs_por_muestra_por_dataset(nombre_dataset, df)

            # 2) recuperar todos los logs acumulados del dataset (en memoria)
            df_salida = cls.obtener_logs_por_muestra_acumulados(nombre_dataset)

        # ───────────── MODO APPEND SOBRE EL ARCHIVO EXTERNO ─────────────
        salida_path = Path(path_salida)
        salida_path.parent.mkdir(parents=True, exist_ok=True)

        if append and salida_path.exists():
            # Leer archivo existente y concatenar
            try:
                df_existente = pd.read_excel(salida_path)
                df_salida = pd.concat([df_existente, df_salida], ignore_index=True)
            except Exception as e:
                print(f"⚠️ No se pudo leer el archivo existente. Se sobrescribirá. Error: {e}")

        # ───────────── EXPORTACIÓN CON FORMATO ─────────────
        with pd.ExcelWriter(str(salida_path), engine="openpyxl") as writer:
            hoja = "Log_Muestras"
            df_salida.to_excel(writer, sheet_name=hoja, index=False)
            ws = writer.sheets[hoja]

            n_rows = len(df_salida)
            n_cols = len(df_salida.columns)
            if n_rows == 0 or n_cols == 0:
                return

            first_row = 2
            last_row = first_row + n_rows - 1
            last_col_letter = get_column_letter(n_cols)

            # Regla 1: is_filtrada == False → rojo suave
            if resaltar_no_filtradas and "is_filtrada" in df_salida.columns:
                col_idx = df_salida.columns.get_loc("is_filtrada") + 1
                col_letter = get_column_letter(col_idx)
                red_fill = PatternFill(start_color="FFF2D7D9",
                                    end_color="FFF2D7D9",
                                    fill_type="solid")
                formula = f"=${col_letter}{first_row}=FALSE"
                ws.conditional_formatting.add(
                    f"A{first_row}:{last_col_letter}{last_row}",
                    FormulaRule(formula=[formula], fill=red_fill)
                )

            # Regla 2 (opcional): resaltar filas por idx_global
            if indices_resaltar is not None and "idx_global" in df_salida.columns:
                s = set(indices_resaltar)
                marcar = df_salida["idx_global"].isin(s)

                if marcar.any():
                    aux_col_idx = n_cols + 1
                    aux_col_letter = get_column_letter(aux_col_idx)
                    ws.cell(row=1, column=aux_col_idx, value="__mark__")

                    for r, v in enumerate(marcar.values, start=2):
                        ws.cell(row=r, column=aux_col_idx, value=bool(v))

                    yellow_fill = PatternFill(start_color="FFFFFFCC",
                                            end_color="FFFFFFCC",
                                            fill_type="solid")
                    formula2 = f"=${aux_col_letter}{first_row}=TRUE"
                    ws.conditional_formatting.add(
                        f"A{first_row}:{last_col_letter}{last_row}",
                        FormulaRule(formula=[formula2], fill=yellow_fill)
                    )

        modo = "append" if append else "overwrite"
        print(f"📄 Log POR MUESTRA guardado en modo {modo}: {salida_path}")



    def exportar_log_muestras_csv(self, path_salida):
        """Exporta el log POR MUESTRA."""
        if not self.logs_por_muestra:
            print("⚠️ No hay log POR MUESTRA para exportar.")
            return
        df = pd.DataFrame(self.logs_por_muestra)

        # Serializar listas en JSON para columnas complejas
        cols_json = (
            "vecinos_all", "clase_vecinos_all", "dist_all",
            "vecinos_min", "dist_vecinos_min"
        )

        for col in cols_json:
            if col in df.columns:
                df[col] = df[col].apply(
                    lambda v: json.dumps(v, ensure_ascii=False)
                    if isinstance(v, (list, tuple, np.ndarray))
                    else v
                )
        df.to_csv(path_salida, index=False)
        print(f"📁 Log por muestra guardado en: {path_salida}")

    # --------------------- Cálculos auxiliares ---------------------
    def _dist(self, A, b):
        metric = getattr(self, "metric", "euclidean")
        if metric == "euclidean":
            if getattr(self, "modo_espacial", "2d") == '3d':
                return np.linalg.norm(A[:, :3] - b[:3], axis=1)
            return np.linalg.norm(A - b, axis=1)
        else:
            # usar el DistanceMetric que ya tenés en PCSMOTE
            if hasattr(self, "_dist_metric"):
                return self._dist_metric.pairwise(A, b.reshape(1, -1))[:, 0]
            # fallback: euclídea
            return np.linalg.norm(A - b, axis=1)

    # --------------------- Logger por muestra ---------------------
    @staticmethod
    def _to_cls_scalar(v):
        """Convierte etiqueta a tipo serializable estable (int si es entero; si no, str)."""
        try:
            arr = np.array(v)
            if np.issubdtype(arr.dtype, np.integer):
                return int(arr.item() if arr.shape == () else v)
        except Exception:
            pass
        try:
            return v.item()
        except Exception:
            return str(v)

    # --- helper local para claves dinámicas seguras ---
    def _kv_with_pct(self, base: str, pct, val):
        """
        Devuelve (clave, valor) donde la clave incluye el percentil.
        - pct=None => sufijo 'none'
        - val=None => se guarda None (no castear a float)
        """
        sufijo = "none" if pct is None else str(int(pct))
        key = f"{base}_{sufijo}"
        value = None if val is None else float(val)
        return key, value

    def _log_muestra(
        self,
        i,                      # índice en X_min
        X, X_min,               # matrices originales y minoritaria
        y,                      # etiquetas globales (para clases de vecinos)
        idxs_min_global,        # mapeo X_min[i] -> índice global en X
        comb,                   # máscara de filtrado por muestra
        riesgo, densidades,     # arrays
        entropias, proporciones_min,  # arrays o None
        pureza_mask, densidad_mask, mask_riesgo,
        umb_ent, umb_den,       # umbrales (float o None)
        vecinos_all_global,     # [n_min, k] índices globales en X
        vecinos_min_global,     # [n_min, k] índices globales minoritarios
        vecinos_validos_counts, # array de conteos válidos por umbral de distancia
        dist_thr_por_muestra,   # array thresholds por muestra
        gen_from_counts,        # dict: idx_global -> sintéticas desde esa semilla
        last_delta_by_seed,     # dict: idx_global -> último delta
        last_neighbor_by_seed,  # dict: idx_global -> último vecino z (idx global)
        idx_local               # índice local usado para densidades (normalmente == i)
    ):
        """
        Registro por muestra con esquema de columnas FIJO.
        - 'percentil_*' almacena el VALOR UMBRAL del percentil (no 25/50/75).
        - Usa el umbral global de distancias para contar vecinos dentro del umbral.
        """

        import numpy as np

        # índice global de la semilla en X
        seed_idx_global = int(idxs_min_global[i])

        # Vecinos globales y minoritarios (índices)
        vecinos_all_lista = []
        for idx in vecinos_all_global[i].tolist():
            vecinos_all_lista.append(int(idx))

        vecinos_min_lista = []
        for idx in vecinos_min_global[i].tolist():
            vecinos_min_lista.append(int(idx))

        # Clases de vecinos_all
        clases_vecinos_all = []
        for idx_vecino in vecinos_all_lista:
            clases_vecinos_all.append(self._to_cls_scalar(y[idx_vecino]))

        # ───────── Meta y umbrales de percentil ─────────
        valor_percentil_dist = None
        valor_percentil_densidad = None
        valor_percentil_entropia = None
        valor_percentil_riesgo = None

        meta = getattr(self, "_meta", {}) if isinstance(getattr(self, "_meta", {}), dict) else {}

        # umbral global de distancias usado en la nueva densidad
        if "umbral_densidad_global" in meta:
            valor_percentil_dist = meta["umbral_densidad_global"]
        elif "valor_percentil_global_elegido" in meta:
            valor_percentil_dist = meta["valor_percentil_global_elegido"]
        elif "valor_percentil_dist" in meta:
            valor_percentil_dist = meta["valor_percentil_dist"]

        if umb_den is not None:
            valor_percentil_densidad = float(umb_den)

        if umb_ent is not None:
            valor_percentil_entropia = float(umb_ent)

        if "umbral_riesgo_min" in meta:
            valor_percentil_riesgo = meta["umbral_riesgo_min"]

        # ───────── Distancias a vecinos y conteo dentro del umbral global ─────────
        xi = X_min[i]

        if len(vecinos_all_lista) > 0:
            # Distancias según la métrica definida en PCSMOTE
            dist_array = self._dist(X[vecinos_all_lista], xi)
        else:
            dist_array = np.array([], dtype=float)

        # Guardar distancias completas solo si se pidió
        if getattr(self, "guardar_distancias", False):
            d_all = dist_array.tolist()
            if len(vecinos_min_lista) > 0:
                d_min = self._dist(X[vecinos_min_lista], xi).tolist()
            else:
                d_min = []
            d_vecinos_min = d_min[:]
        else:
            d_all = None
            d_vecinos_min = None

        # Conteo de vecinos dentro del umbral global de distancia
        if dist_array.size > 0 and valor_percentil_dist is not None:
            mascara_en_umbral = dist_array <= float(valor_percentil_dist)
            cant_vecinos_en_p_i = int(np.sum(mascara_en_umbral))

            cant_min_en_p_i = 0
            for idx_vecino, en_umbral in zip(vecinos_all_lista, mascara_en_umbral):
                if en_umbral and y[idx_vecino] == 1:
                    cant_min_en_p_i += 1
        else:
            cant_vecinos_en_p_i = 0
            cant_min_en_p_i = 0

        # ───────── Construcción explícita del registro ─────────
        registro = {
            "idx_global": seed_idx_global,
            "clase_objetivo": None,  # se pisa desde fit_resample_multiclass si corresponde
            "es_semilla_valida": bool(comb[i]),
            "k": int(getattr(self, "k", 0)),

            # percentiles usados (VALOR DE UMBRAL, no 25/50/75)
            "valor_percentil_dist": valor_percentil_dist,
            "valor_umbral_densidad": valor_percentil_densidad,
            "valor_percentil_entropia": valor_percentil_entropia,
            "valor_percentil_riesgo": valor_percentil_riesgo,

            # umbrales asociados
            "umbral_densidad": None if umb_den is None else float(umb_den),
            "umbral_entropia": None if umb_ent is None else float(umb_ent),

            # métricas locales
            "criterio_pureza": getattr(self, "criterio_pureza", None),
            "riesgo": float(riesgo[i]),
            "densidad": float(densidades[idx_local]),
            "entropia": None if entropias is None else float(entropias[i]),
            "proporcion_min": None if proporciones_min is None else float(proporciones_min[i]),
            "pasa_pureza": bool(pureza_mask[i]),
            "pasa_densidad": bool(densidad_mask[i]),
            "pasa_riesgo": bool(mask_riesgo[i]),

            # diagnóstico de threshold de distancia por muestra (sigue como antes)
            "vecinos_validos_por_percentil": int(vecinos_validos_counts[i]),
            "thr_dist_percentil": float(dist_thr_por_muestra[i]),

            # uso en síntesis
            "synthetics_from_this_seed": int(gen_from_counts.get(seed_idx_global, 0)),
            "last_delta": last_delta_by_seed.get(seed_idx_global, None),
            "last_neighbor_z": last_neighbor_by_seed.get(seed_idx_global, None),
        }

        # ---- NUEVAS CABECERAS DEL NUEVO SISTEMA DE DENSIDAD (por muestra, no promedio) ----
        registro["percentil_densidad_distancias_elegido"] = meta.get("percentil_densidad_distancias_elegido")
        registro["valor_percentil_global_elegido"] = meta.get("valor_percentil_global_elegido")
        registro["k_global"] = meta.get("k_global")
        registro["cant_vecinos_en_p_elegido"] = cant_vecinos_en_p_i
        registro["cant_min_en_p_elegido"] = cant_min_en_p_i

        self.logs_por_muestra.append(registro)
